from fpdf import FPDF
import qrcode
import openpyxl
from pathlib import Path
import os

def generate_qr_code(data, output_filename="my_qr_code.png"):

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=2,
        border=1,
    )
    qr.add_data(data)
    qr.make(fit=True)

    # Create QR code image
    qr_image = qr.make_image(fill_color="black", back_color="white")

    # Save the image
    qr_image.save(output_filename)


def delete_png_files_based_on_excel_column(excel_path):

    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # Assuming the column index is fixed (e.g., column B)
        column_index = 3  # Adjust as needed

        # Get the folder path where the Excel file is located
        folder_path = os.path.dirname(excel_path)

        # Iterate through the values in the specified column
        for cell in sheet.iter_rows(min_row=2, min_col=column_index, values_only=True):
            value = cell[0]
            if value:
                # Construct the PNG file path
                png_file_path = os.path.join(folder_path, f"{value}.png")

                # Check if the PNG file exists and delete it
                if os.path.exists(png_file_path):
                    os.remove(png_file_path)

    except Exception as e:
        print(f"An error occurred: {str(e)}")

        

def read_excel_data(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]  # Replace with your sheet name
    data = []

    for row in sheet.iter_rows(min_row=2):  # Skip header row
        eid, personnel, comment, ruta, pcType = row[2], row[1], row[28], row[31], row[0]  # Adjust column indices #comment->27 #ruta -> 30
        data.append((eid.value, personnel.value, comment.value, ruta.value, pcType.value))

    return data

def file_exists_using_pathlib(file_path: str) -> bool:
    my_file = Path(file_path)
    return my_file.is_file()
def create_labels(data, excelName):
    # Create an FPDF object
    pdf = FPDF()
    pdf.add_page()

    # Set font (Arial, bold, size 16)
    cantUsers = len(data)
    loopsQuant = (cantUsers// 8) + 1
    it = 0
    i = 0
    while it <= loopsQuant:


        # Define label dimensions (adjust as needed)
        label_width = 80
        label_height = 45
        label_spacing = 15 # Space between labels
        labels_per_page = 8
        label_count = 0
        
        for row in range(4):
            for col in range(2):
                if i < cantUsers:
                # Extract data for the label (EID, Personnel, Comment, Ruta)
                    eid, personnel, comment, ruta, pcType = data[i]  # Adjust based on your data structure

                    # Generate QR code (use your existing function)
                    generate_qr_code(eid, output_filename=str(eid) + ".png")

                    x = col * (label_width + label_spacing)
                    y = row * (label_height + label_spacing)

                    # Create label content (adjust positions as needed)
                    pdf.set_font('Arial', size=10)
                    pdf.set_xy(x+8, y+20)
                    pdf.multi_cell(label_width + 5, label_height - 5, txt="__________________________________________", border=1)

                    pdf.set_font('Arial', "B", 13)
                    pdf.set_xy(x+8, y+27)
                    pdf.multi_cell(label_width + 5, label_height - 5, txt=f"EID: {eid}", border=0)

                    # Add QR code image (adjust position)
                    pdf.image(str(eid) + ".png", x=x + 75, y=y + 22, w=15, h=15)
                    

                    pdf.set_font('Arial', "B", 10) #PERSONNEL
                    pdf.set_xy(x+8, y+20)
                    pdf.multi_cell(label_width, 17, txt=f"Personnel N°: {personnel}", border=0)

                    pdf.set_font('Arial', "B", 13)
                    pdf.set_xy(x+8, y+18) #EID 
                    pdf.multi_cell(label_width, 10, txt=str(eid) , border=0)

                    pdf.set_font('Arial', size=10)
                    pdf.set_xy(x+8, y+22) #COMMENT / COST CENTER
                    pdf.multi_cell(label_width, 24, txt= f"Comment: {comment}" , border=0)


                    pdf.set_font('Arial', size=9)
                    pdf.set_xy(x+8, y+27) #COMMENT / COST CENTER
                    pdf.multi_cell(label_width, 24, txt= pcType, border=0)                










                    pdf.set_font('Arial', 'B', 10)
                    pdf.set_xy(x+8, y+11) #RUTA
                    pdf.multi_cell(label_width + 15, 85, f"Ruta: {ruta}", border=0)

                    label_count += 1
                    if label_count >= labels_per_page:
                        pdf.add_page()  # Start a new page
                        label_count = 0
                    i += 1
        it += 1
    # Save the PDF
    pdf.output("Etiquetas " + str(excelName)+'.pdf', 'F')


def separate_excel_files(root_file_path: str) -> None:


    script_directory = os.path.dirname(os.path.realpath(__file__))
    filePath = os.path.join(script_directory, root_file_path)

    if file_exists_using_pathlib(filePath):
        root_workbook = openpyxl.load_workbook(root_file_path)
        root_sheet = root_workbook.active

        # Get unique values from the specified column (e.g., column F)
        unique_values = set(root_sheet["AF2:AF" + str(root_sheet.max_row)])

        # Create separate Excel files for each unique value
        for value_cell in unique_values:
            value = value_cell[0].value
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active

            # Copy the header row from the root sheet
            header_row = root_sheet[1]
            for cell in header_row:
                new_sheet[cell.column_letter + "1"] = cell.value

            # Copy matching rows to the new sheet
            for row in root_sheet.iter_rows(min_row=2, values_only=True):
                if row[31] == value:  # Assuming column AE contains the values to match
                    new_sheet.append(row)

            # Save the new workbook
            new_workbook.save(f"{value}.xlsx")

        # Close the root workbook
        root_workbook.close()
        print("Separate Excel files created and matching rows copied successfully!")
    else:
        print(f"The file {filePath} does not exist")


listOfPossibleRoutes = ['ZzZZZzzzZ', 'xXxXxX + yyYyY']

fileName = input("Enter the excel name: \n")

root_excel_file = fileName + ".xlsx"



separate_excel_files(root_excel_file)


for route in listOfPossibleRoutes:
    if file_exists_using_pathlib(file_path=str(route)+'.xlsx'):
        create_labels(read_excel_data(filename=str(route)+'.xlsx'), str(route))
    else:
        print(str(route) + " does not exists in this filer\n")

delete_png_files_based_on_excel_column(root_excel_file)